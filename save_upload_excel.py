from RequestBaseManager import RequestBaseManager
from io import BytesIO
from datetime import datetime
from stock import Stock
from openpyxl import load_workbook
import tokenManager
import storemgr
import json

def getItems(ws) -> set:
    
    results = set()
    
    for i in range(1, ws.max_row):
    
        obj = Stock()
        
        index = str(i + 1)

        obj.stockId = ws['A' + index].value
        
        temp = str(ws['B' + index].value)

        t = datetime.strptime(temp[0:10], '%Y-%m-%d')
        
        obj.datestr = t.strftime('%Y-%m-%d')

        obj.increase = str(ws['C' + index].value)

        results.add(obj)
        
    return results

class SaveStocks(RequestBaseManager):
    
    def post (self, *args, **kwargs):

        token = self.get_body_argument('token')

        msg = 'ok'

        if tokenManager.instance().checkToken(token) is not True:

            msg = 'token出错'

        try:
            data = self.request.files['upload'][0].body

            wb = load_workbook(filename = BytesIO(data))

            ws = wb.active

            storemgr.saveStocks(getItems(ws))

        except:
            msg = '上传文件出错'

        if msg == 'ok':

            self.write({'success': 1})

        else:

            self.write({'success': 0, 'msg':msg})